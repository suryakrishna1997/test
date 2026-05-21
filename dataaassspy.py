import pandas as pd
import tkinter as tk
from tkinter import messagebox
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_multi_sheet_excel():
    file_name = "Test_Case_Template_MultiSheet.xlsx"
    
    columns = [
        "Sr. No.", "SCREEN_ID", "REQUIREMENT_ID", "KPIT_TC_ID", "Name", 
        "Objective", "Precondition", "Test Script (Step-by-Step) - Step", 
        "Test Script (Step-by-Step) - Test Data", "Test Script (Step-by-Step) - Expected Result", 
        "FEATURE_1", "REFERENCE_DOC", "Labels", "ANDROID_VER", "VEHICLE_MODEL", 
        "REGION", "VARIANT_CODE", "TEST_LEVEL", "TEST_TYPE", "HW_REQUIRED", 
        "SRL_VERSION", "SUB_ID", "Status", "Priority", "Owner", "Folder", 
        "Vehicle model required for Execution\ne.g. 30AA Q", "TC creator", 
        "L1 Review by", "L1 Review Comments", "TC Status", "L2 Review by", 
        "L2 Review Comments", "TC Status", "L3 Review by", "L3 Review Comments", 
        "TC Status", "Remarks", "Execution feasibility Result (Pass/Fail)", 
        "Execution Model", "Execution Region", "Executed by"
    ]

    try:
        wb = Workbook()
        # Remove the default sheet created by Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        # --- Styles Definition ---
        # Color: 99FF33
        header_fill = PatternFill(start_color="99FF33", end_color="99FF33", fill_type="solid")
        header_font = Font(color="000000", bold=True, size=11)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Border Styles (Thin lines for all sides)
        thin_side = Side(border_style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # --- Loop to create 10 sheets ---
        for i in range(1, 11):
            sheet_name = f"Sheet_{i}"
            ws = wb.create_sheet(title=sheet_name)

            # 1. Add Headers
            ws.append(columns)

            # 2. Add 25 empty rows (to make it exactly 25 rows of data + 1 header row)
            # If you want 25 rows TOTAL including header, change range to 25
            for _ in range(25):
                ws.append([""] * len(columns))

            # 3. Apply Styling and Borders to the used range
            # We iterate through the cells that were just created
            for row in ws.iter_rows(min_row=1, max_row=26, min_col=1, max_col=len(columns)):
                for cell in row:
                    cell.border = border  # Apply border to every cell
                    cell.alignment = alignment # Center everything
                    
                    # If it's the first row, apply header color/font
                    if cell.row == 1:
                        cell.fill = header_fill
                        cell.font = header_font

            # 4. Auto-adjust Column Widths (Optimized for 10 sheets)
            for col in ws.columns:
                column_letter = col[0].column_letter
                # Set a standard width for these specific columns to keep file size manageable
                ws.column_dimensions[column_letter].width = 20

        # Save the file
        wb.save(file_name)
        messagebox.showinfo("Success", f"10 Sheets generated successfully!\nFile: {os.path.abspath(file_name)}")
    
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# --- GUI Setup ---
root = tk.Tk()
root.title("KPIT Excel Pro")
root.geometry("350x200")
root.configure(bg="#f0f0f0")

label = tk.Label(root, text="Multi-Sheet Template Generator", font=("Arial", 12, "bold"), bg="#f0f0f0", pady=20)
label.pack()

btn = tk.Button(root, text="Generate 10 Sheets (25 Rows Each)", command=create_multi_sheet_excel, 
                bg="#99FF33", fg="black", font=("Arial", 10, "bold"), 
                padx=20, pady=10, cursor="hand2")
btn.pack()

root.mainloop()

# This is AI generated code, please refer KPIT AI Policy before using this in your projects
